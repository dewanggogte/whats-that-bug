#!/usr/bin/env python3
"""
What's That Bug — Analytics Pipeline

Pulls data from Umami Cloud API + Google Sheets xlsx export.
Outputs CSVs and a self-contained HTML dashboard.

Usage:
    python dump.py                              # Dump + dashboard, last 30 days
    python dump.py --days 90                    # Last 90 days
    python dump.py --start 2026-04-01           # Custom start (end defaults to today)
    python dump.py --xlsx ../feedback.xlsx      # Include Google Sheets game data
    python dump.py --dump-only                  # CSVs only, skip dashboard
    python dump.py --dashboard-only             # Dashboard from existing CSVs

Requires UMAMI_API_KEY and UMAMI_WEBSITE_ID in .env
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from urllib.parse import urlencode

SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
UMAMI_DIR = OUTPUT_DIR / "umami"
SHEETS_DIR = OUTPUT_DIR / "sheets"


# ---------------------------------------------------------------------------
# .env loader (no external dependency)
# ---------------------------------------------------------------------------

def load_env():
    env_file = SCRIPT_DIR / ".env"
    if not env_file.exists():
        return
    with open(env_file) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            key, _, value = line.partition("=")
            key, value = key.strip(), value.strip()
            if value and key:
                os.environ.setdefault(key, value)


# ---------------------------------------------------------------------------
# Umami Cloud API client
# ---------------------------------------------------------------------------

class UmamiClient:
    """Thin wrapper around the Umami Cloud REST API (v2)."""

    BASE_URL = "https://api.umami.is"

    def __init__(self, api_key: str, website_id: str):
        self.api_key = api_key
        self.website_id = website_id

    def _get(self, path: str, params: dict | None = None) -> dict | list:
        url = f"{self.BASE_URL}{path}"
        if params:
            url += "?" + urlencode(params)
        req = Request(url)
        req.add_header("x-umami-api-key", self.api_key)
        req.add_header("Accept", "application/json")
        try:
            with urlopen(req, timeout=30) as resp:
                return json.loads(resp.read())
        except HTTPError as e:
            body = e.read().decode(errors="replace")
            print(f"  [ERROR] Umami API {e.code}: {body[:300]}", file=sys.stderr)
            raise
        except URLError as e:
            print(f"  [ERROR] Network error: {e.reason}", file=sys.stderr)
            raise

    def stats(self, start_ms: int, end_ms: int) -> dict:
        return self._get(f"/api/websites/{self.website_id}/stats", {
            "startAt": start_ms, "endAt": end_ms,
        })

    def pageviews(self, start_ms: int, end_ms: int, unit: str = "day") -> dict:
        return self._get(f"/api/websites/{self.website_id}/pageviews", {
            "startAt": start_ms, "endAt": end_ms, "unit": unit,
        })

    def metrics(self, start_ms: int, end_ms: int, metric_type: str) -> list:
        return self._get(f"/api/websites/{self.website_id}/metrics", {
            "startAt": start_ms, "endAt": end_ms, "type": metric_type,
        })

    def events(self, start_ms: int, end_ms: int, unit: str = "day") -> list:
        return self._get(f"/api/websites/{self.website_id}/events", {
            "startAt": start_ms, "endAt": end_ms, "unit": unit,
        })


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def write_csv(path: Path, rows: list[dict], fieldnames: list[str] | None = None):
    if not rows:
        print(f"  (no data for {path.name})")
        return
    if fieldnames is None:
        fieldnames = list(rows[0].keys())
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"  {path.name}: {len(rows)} rows")


def read_csv(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Umami dump
# ---------------------------------------------------------------------------

def dump_umami(api_key: str, website_id: str, start_ms: int, end_ms: int):
    print("\n--- Umami Cloud ---")
    client = UmamiClient(api_key, website_id)
    UMAMI_DIR.mkdir(parents=True, exist_ok=True)

    # Summary stats
    print("  Fetching stats...")
    stats = client.stats(start_ms, end_ms)
    write_csv(UMAMI_DIR / "stats.csv", [flatten_stats(stats)])

    # Pageviews time series
    print("  Fetching pageviews...")
    pv = client.pageviews(start_ms, end_ms, unit="day")
    pv_rows = [{"date": p["x"], "pageviews": p["y"]} for p in pv.get("pageviews", [])]
    sess_rows = [{"date": s["x"], "sessions": s["y"]} for s in pv.get("sessions", [])]
    # Merge pageviews and sessions by date
    sessions_map = {r["date"]: r["sessions"] for r in sess_rows}
    for row in pv_rows:
        row["sessions"] = sessions_map.get(row["date"], 0)
    write_csv(UMAMI_DIR / "pageviews.csv", pv_rows)

    # Metrics (with small delay between calls to be respectful)
    metric_types = ["url", "referrer", "browser", "os", "device", "country"]
    file_names = ["pages", "referrers", "browsers", "os", "devices", "countries"]
    for mtype, fname in zip(metric_types, file_names):
        print(f"  Fetching {fname}...")
        data = client.metrics(start_ms, end_ms, mtype)
        rows = [{"name": item["x"], "count": item["y"]} for item in data]
        write_csv(UMAMI_DIR / f"{fname}.csv", rows)
        time.sleep(0.3)  # rate-limit courtesy

    # Events
    print("  Fetching events...")
    try:
        ev = client.events(start_ms, end_ms, unit="day")
        ev_rows = [{"event": item["x"], "date": item["t"], "count": item["y"]} for item in ev]
        write_csv(UMAMI_DIR / "events.csv", ev_rows)
    except Exception as e:
        print(f"  (events endpoint failed: {e} — skipping)")

    print("  Done.")


def flatten_stats(stats: dict) -> dict:
    """Flatten nested {value, change} objects into a single row."""
    flat = {}
    for key, val in stats.items():
        if isinstance(val, dict):
            flat[key] = val.get("value", val)
            change = val.get("change")
            if change is not None:
                flat[f"{key}_change"] = change
        else:
            flat[key] = val
    return flat


# ---------------------------------------------------------------------------
# Google Sheets xlsx dump
# ---------------------------------------------------------------------------

def dump_sheets(xlsx_path: Path):
    print("\n--- Google Sheets ---")
    try:
        import openpyxl
    except ImportError:
        print("  [ERROR] openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return

    if not xlsx_path.exists():
        print(f"  [ERROR] File not found: {xlsx_path}", file=sys.stderr)
        return

    SHEETS_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            print(f"  {sheet_name}: empty, skipping")
            continue

        # Use first row as headers, clean up None columns
        raw_headers = rows[0]
        headers = [h for h in raw_headers if h is not None]
        n_cols = len(headers)

        data = []
        for row in rows[1:]:
            values = row[:n_cols]
            data.append({h: (str(v) if v is not None else "") for h, v in zip(headers, values)})

        fname = sheet_name.lower().replace(" ", "_") + ".csv"
        write_csv(SHEETS_DIR / fname, data, fieldnames=headers)

    wb.close()
    print("  Done.")


# ---------------------------------------------------------------------------
# Dashboard HTML generator
# ---------------------------------------------------------------------------

def generate_dashboard(start_date: str, end_date: str):
    print("\n--- Dashboard ---")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load Umami CSVs (already small/aggregated)
    data: dict = {}
    if UMAMI_DIR.exists():
        for csv_file in UMAMI_DIR.glob("*.csv"):
            data[f"umami/{csv_file.stem}"] = read_csv(csv_file)

    # Load and PRE-AGGREGATE sheet data (avoid embedding 20K raw rows)
    events = read_csv(SHEETS_DIR / "events.csv") if (SHEETS_DIR / "events.csv").exists() else []
    leaderboard = read_csv(SHEETS_DIR / "leaderboard.csv") if (SHEETS_DIR / "leaderboard.csv").exists() else []

    if events:
        data["game/aggregated"] = aggregate_game_events(events)
    if leaderboard:
        data["game/leaderboard_count"] = [{"count": len(leaderboard)}]

    if not data:
        print("  [ERROR] No CSV data found. Run --dump first.", file=sys.stderr)
        return

    data_json = json.dumps(data, default=str)
    print(f"  Dashboard payload: {len(data_json) // 1024}KB")

    html = build_dashboard_html(data_json, start_date, end_date)
    out_path = OUTPUT_DIR / "dashboard.html"
    out_path.write_text(html)
    print(f"  Generated: {out_path}")
    print(f"  Open in browser: file://{out_path.resolve()}")


def aggregate_game_events(events: list[dict]) -> dict:
    """Pre-compute all aggregations the dashboard needs from raw events."""
    agg: dict = {
        "total_sessions": 0,
        "total_completes": 0,
        "total_daily_starts": 0,
        "total_daily_completes": 0,
        "sessions_by_day": {},
        "sessions_by_set": {},
        "scores_by_set": {},   # {set: {total, count}}
        "daily_by_day": {},    # {day: {starts, completes}}
        "event_types": {},
        # Feedback aggregations
        "difficulty_ratings": {},       # {rating: count}
        "play_again_intent": {},        # {yes/maybe/no: count}
        "feedback_texts": [],           # free text feedback entries
        "bad_photo_reports": {},        # {observation_id: {count, species}}
        "round_reactions": {},          # {round_num: {too_easy: N, just_right: N, too_hard: N}}
        "observation_miss_data": {},    # {obs_id: {total, wrong, times, species, set}}
        "session_round_counts": {},     # {session_id: max_round_num}
    }

    for e in events:
        etype = e.get("type", "")
        day = (e.get("timestamp") or "")[:10]

        # Event type counts
        if etype:
            agg["event_types"][etype] = agg["event_types"].get(etype, 0) + 1

        if etype == "session_start":
            agg["total_sessions"] += 1
            if day:
                agg["sessions_by_day"][day] = agg["sessions_by_day"].get(day, 0) + 1
            s = e.get("set", "Unknown")
            agg["sessions_by_set"][s] = agg["sessions_by_set"].get(s, 0) + 1

        elif etype == "session_end":
            agg["total_completes"] += 1

        elif etype == "round_complete":
            data_json = e.get("data_json", "")
            if data_json:
                try:
                    d = json.loads(data_json)
                    if not isinstance(d, dict):
                        continue
                    s = d.get("set") or e.get("set", "Unknown")
                    if s not in agg["scores_by_set"]:
                        agg["scores_by_set"][s] = {"total": 0, "count": 0}
                    agg["scores_by_set"][s]["total"] += d.get("score", 0)
                    agg["scores_by_set"][s]["count"] += 1

                    # --- Image quality / round-level tracking ---
                    obs_id = str(d.get("observation_id", ""))
                    if obs_id:
                        if obs_id not in agg["observation_miss_data"]:
                            agg["observation_miss_data"][obs_id] = {
                                "total": 0, "wrong": 0, "times": [],
                                "species": d.get("correct_answer", ""),
                                "set": s,
                            }
                        entry = agg["observation_miss_data"][obs_id]
                        entry["total"] += 1
                        if d.get("score", 0) == 0:
                            entry["wrong"] += 1
                        time_ms = d.get("time_taken_ms")
                        if time_ms is not None:
                            try:
                                entry["times"].append(int(time_ms))
                            except (ValueError, TypeError):
                                pass

                    # Track max round per session for drop-off analysis
                    sid = d.get("session_id", "")
                    rnd = d.get("round")
                    if sid and rnd is not None:
                        try:
                            rnd_int = int(rnd)
                            cur = agg["session_round_counts"].get(sid, 0)
                            if rnd_int > cur:
                                agg["session_round_counts"][sid] = rnd_int
                        except (ValueError, TypeError):
                            pass
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass

        elif etype == "daily_start":
            agg["total_daily_starts"] += 1
            if day:
                if day not in agg["daily_by_day"]:
                    agg["daily_by_day"][day] = {"starts": 0, "completes": 0}
                agg["daily_by_day"][day]["starts"] += 1

        elif etype == "daily_complete":
            agg["total_daily_completes"] += 1
            if day:
                if day not in agg["daily_by_day"]:
                    agg["daily_by_day"][day] = {"starts": 0, "completes": 0}
                agg["daily_by_day"][day]["completes"] += 1

        elif etype == "session_feedback":
            data_json = e.get("data_json", "")
            if data_json:
                try:
                    d = json.loads(data_json)
                    if not isinstance(d, dict):
                        continue
                    # Difficulty rating
                    rating = d.get("difficulty_rating", "")
                    if rating:
                        agg["difficulty_ratings"][str(rating)] = agg["difficulty_ratings"].get(str(rating), 0) + 1
                    # Play again intent
                    play_again = d.get("play_again", "")
                    if play_again:
                        agg["play_again_intent"][play_again] = agg["play_again_intent"].get(play_again, 0) + 1
                    # Free text feedback
                    free_text = (d.get("free_text") or "").strip()
                    if free_text:
                        agg["feedback_texts"].append({
                            "text": free_text,
                            "difficulty": rating,
                            "play_again": play_again,
                            "timestamp": d.get("timestamp", "")[:10],
                        })
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass

        elif etype == "round_reaction":
            data_json = e.get("data_json", "")
            if data_json:
                try:
                    d = json.loads(data_json)
                    if not isinstance(d, dict):
                        continue
                    rnd = d.get("round")
                    # Field is called "difficulty" in the actual data
                    reaction = d.get("difficulty", "")
                    if rnd is not None and reaction:
                        rnd_key = str(int(rnd))
                        if rnd_key not in agg["round_reactions"]:
                            agg["round_reactions"][rnd_key] = {"too_easy": 0, "just_right": 0, "too_hard": 0}
                        if reaction in agg["round_reactions"][rnd_key]:
                            agg["round_reactions"][rnd_key][reaction] += 1
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass

        elif etype == "bad_photo":
            data_json = e.get("data_json", "")
            if data_json:
                try:
                    d = json.loads(data_json)
                    if not isinstance(d, dict):
                        continue
                    obs_id = str(d.get("observation_id", ""))
                    if obs_id:
                        if obs_id not in agg["bad_photo_reports"]:
                            agg["bad_photo_reports"][obs_id] = {"count": 0, "species": d.get("species", "")}
                        agg["bad_photo_reports"][obs_id]["count"] += 1
                except (json.JSONDecodeError, TypeError, ValueError):
                    pass

    # Compute average scores
    avg_scores = {}
    for s, v in agg["scores_by_set"].items():
        avg_scores[s] = round(v["total"] / v["count"]) if v["count"] > 0 else 0
    agg["avg_scores_by_set"] = avg_scores

    # --- Derived stats ---

    # Worst observations by miss rate (min 3 attempts), top 50
    worst_obs = []
    for obs_id, info in agg["observation_miss_data"].items():
        if info["total"] >= 3:
            miss_rate = round(info["wrong"] / info["total"] * 100, 1)
            avg_time = round(sum(info["times"]) / len(info["times"])) if info["times"] else 0
            bad_reports = agg["bad_photo_reports"].get(obs_id, {}).get("count", 0)
            worst_obs.append({
                "obs_id": obs_id,
                "species": info["species"],
                "miss_rate": miss_rate,
                "attempts": info["total"],
                "wrong": info["wrong"],
                "avg_time_ms": avg_time,
                "bad_photo_reports": bad_reports,
                "set": info["set"],
            })
    worst_obs.sort(key=lambda x: x["miss_rate"], reverse=True)
    agg["worst_observations"] = worst_obs[:50]

    # Confusion pairs: top 20 (correct_answer, user_answer) where score=0
    confusion_counts: dict[tuple, int] = {}
    for e in events:
        if e.get("type") != "round_complete":
            continue
        data_json = e.get("data_json", "")
        if not data_json:
            continue
        try:
            d = json.loads(data_json)
            if not isinstance(d, dict):
                continue
            if d.get("score", 0) == 0:
                correct = d.get("correct_answer", "Unknown")
                picked = d.get("user_answer", "Unknown")
                pair = (correct, picked)
                confusion_counts[pair] = confusion_counts.get(pair, 0) + 1
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    confusion_sorted = sorted(confusion_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    agg["confusion_pairs"] = [
        {"correct": p[0], "picked": p[1], "count": c}
        for (p, c) in confusion_sorted
    ]

    # Time anomalies: observations where avg time > 2x global median
    all_times = []
    for info in agg["observation_miss_data"].values():
        all_times.extend(info["times"])
    if all_times:
        all_times_sorted = sorted(all_times)
        mid = len(all_times_sorted) // 2
        global_median = (all_times_sorted[mid] + all_times_sorted[~mid]) / 2
        time_anomalies = []
        for obs_id, info in agg["observation_miss_data"].items():
            if info["times"]:
                avg_t = sum(info["times"]) / len(info["times"])
                if global_median > 0 and avg_t > 2 * global_median:
                    time_anomalies.append({
                        "obs_id": obs_id,
                        "species": info["species"],
                        "avg_time_ms": round(avg_t),
                        "median_ratio": round(avg_t / global_median, 1),
                        "attempts": info["total"],
                    })
        time_anomalies.sort(key=lambda x: x["median_ratio"], reverse=True)
        agg["time_anomalies"] = time_anomalies[:50]
        agg["global_median_time_ms"] = round(global_median)
    else:
        agg["time_anomalies"] = []
        agg["global_median_time_ms"] = 0

    # Drop-off by round: what % of sessions reach each round (1-10)
    total_tracked = len(agg["session_round_counts"])
    dropoff = {}
    if total_tracked > 0:
        for rnd in range(1, 11):
            reached = sum(1 for mx in agg["session_round_counts"].values() if mx >= rnd)
            dropoff[str(rnd)] = round(reached / total_tracked * 100, 1)
    agg["dropoff_by_round"] = dropoff

    # Round accuracy by set: per-round accuracy aggregated by set
    # {set: {round: {correct: N, total: N}}}
    round_acc_by_set: dict[str, dict[str, dict]] = {}
    for e in events:
        if e.get("type") != "round_complete":
            continue
        data_json = e.get("data_json", "")
        if not data_json:
            continue
        try:
            d = json.loads(data_json)
            if not isinstance(d, dict):
                continue
            s = d.get("set") or e.get("set", "Unknown")
            rnd = str(int(d.get("round", 0)))
            if s not in round_acc_by_set:
                round_acc_by_set[s] = {}
            if rnd not in round_acc_by_set[s]:
                round_acc_by_set[s][rnd] = {"correct": 0, "total": 0}
            round_acc_by_set[s][rnd]["total"] += 1
            if d.get("score", 0) > 0:
                round_acc_by_set[s][rnd]["correct"] += 1
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    agg["round_accuracy_by_set"] = round_acc_by_set

    # --- Insights ---
    insights = []

    # Red: observations with >80% miss rate and >=5 attempts
    critical_obs = [o for o in worst_obs if o["miss_rate"] > 80 and o["attempts"] >= 5]
    for o in critical_obs[:5]:
        insights.append({
            "type": "warning",
            "text": f"Observation {o['obs_id']} ({o['species']}) has {o['miss_rate']}% miss rate over {o['attempts']} attempts"
            + (f" and {o['bad_photo_reports']} bad photo reports" if o["bad_photo_reports"] else "")
            + f" — consider replacing.",
        })

    # Yellow: sets with >30% "too hard" reactions
    reactions_by_set: dict[str, dict] = {}
    for e in events:
        if e.get("type") != "round_reaction":
            continue
        data_json = e.get("data_json", "")
        if not data_json:
            continue
        try:
            d = json.loads(data_json)
            if not isinstance(d, dict):
                continue
            s = d.get("set", "Unknown")
            reaction = d.get("difficulty", "")
            if s not in reactions_by_set:
                reactions_by_set[s] = {"too_easy": 0, "just_right": 0, "too_hard": 0, "total": 0}
            if reaction in ("too_easy", "just_right", "too_hard"):
                reactions_by_set[s][reaction] += 1
                reactions_by_set[s]["total"] += 1
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    for s, counts in reactions_by_set.items():
        if counts["total"] >= 5:
            hard_pct = round(counts["too_hard"] / counts["total"] * 100, 1)
            if hard_pct > 30:
                insights.append({
                    "type": "info",
                    "text": f'Set "{s}" has {hard_pct}% "too hard" reactions ({counts["too_hard"]}/{counts["total"]} rounds) — review difficulty balance.',
                })

    # Green: play-again rate
    play_again = agg["play_again_intent"]
    total_responses = sum(play_again.values())
    if total_responses >= 3:
        yes_count = play_again.get("yes", 0)
        yes_pct = round(yes_count / total_responses * 100, 1)
        if yes_pct >= 60:
            insights.append({
                "type": "positive",
                "text": f"{yes_pct}% of players ({yes_count}/{total_responses}) say they'd play again!",
            })

    # Green: high completion rate
    if agg["total_sessions"] > 0:
        comp_pct = round(agg["total_completes"] / agg["total_sessions"] * 100, 1)
        if comp_pct >= 50:
            insights.append({
                "type": "positive",
                "text": f"{comp_pct}% game completion rate ({agg['total_completes']}/{agg['total_sessions']} sessions).",
            })

    agg["insights"] = insights

    # Clean up large intermediate data not needed in the dashboard JSON
    del agg["observation_miss_data"]
    del agg["session_round_counts"]

    return agg


def build_dashboard_html(data_json: str, start_date: str, end_date: str) -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>What's That Bug — Analytics</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
  :root {{
    --bg: #FDF6F0;
    --surface: #FFFFFF;
    --border: #E8DDD4;
    --text: #3D2B1F;
    --text-muted: #8B7355;
    --primary: #C45D3E;
    --primary-light: #E8A87C;
    --accent: #85603F;
    --success: #5B8C5A;
    --shadow: 0 1px 3px rgba(61,43,31,0.08);
  }}
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif;
    background: var(--bg); color: var(--text);
    line-height: 1.5; padding: 24px; max-width: 1200px; margin: 0 auto;
  }}
  header {{
    display: flex; justify-content: space-between; align-items: baseline;
    border-bottom: 2px solid var(--primary); padding-bottom: 12px; margin-bottom: 24px;
  }}
  header h1 {{ font-size: 1.5rem; color: var(--primary); }}
  header .date-range {{ color: var(--text-muted); font-size: 0.9rem; }}

  /* Summary cards */
  .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 16px; margin-bottom: 32px; }}
  .card {{
    background: var(--surface); border: 1px solid var(--border); border-radius: 8px;
    padding: 16px; box-shadow: var(--shadow);
  }}
  .card .label {{ font-size: 0.75rem; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; }}
  .card .value {{ font-size: 1.8rem; font-weight: 700; color: var(--primary); margin-top: 4px; }}
  .card .change {{ font-size: 0.8rem; margin-top: 2px; }}
  .card .change.up {{ color: var(--success); }}
  .card .change.down {{ color: var(--primary); }}

  /* Sections */
  section {{ margin-bottom: 40px; }}
  section h2 {{
    font-size: 1.1rem; color: var(--accent); margin-bottom: 16px;
    padding-bottom: 6px; border-bottom: 1px solid var(--border);
  }}
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
  .grid-3 {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; }}
  @media (max-width: 768px) {{
    .grid, .grid-3 {{ grid-template-columns: 1fr; }}
  }}
  .chart-box {{
    background: var(--surface); border: 1px solid var(--border); border-radius: 8px;
    padding: 16px; box-shadow: var(--shadow);
  }}
  .chart-box h3 {{ font-size: 0.85rem; color: var(--text-muted); margin-bottom: 12px; }}
  .chart-box canvas {{ width: 100% !important; }}

  /* Tables */
  .data-table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
  .data-table th {{ text-align: left; color: var(--text-muted); font-weight: 500; padding: 6px 8px; border-bottom: 1px solid var(--border); }}
  .data-table td {{ padding: 6px 8px; border-bottom: 1px solid var(--border); }}
  .data-table tr:last-child td {{ border-bottom: none; }}
  .bar {{ display: inline-block; height: 8px; border-radius: 4px; background: var(--primary-light); }}

  .empty-state {{ color: var(--text-muted); font-style: italic; padding: 24px; text-align: center; }}

  /* Insights panel */
  .insights {{ margin-bottom: 24px; }}
  .insights h2 {{
    font-size: 1.1rem; color: var(--accent); margin-bottom: 12px;
    padding-bottom: 6px; border-bottom: 1px solid var(--border);
  }}
  .insight-cards {{ display: flex; flex-wrap: wrap; gap: 12px; }}
  .insight-card {{
    flex: 1 1 300px; padding: 12px 16px; border-radius: 8px;
    font-size: 0.85rem; line-height: 1.5; border-left: 4px solid;
    background: var(--surface); box-shadow: var(--shadow);
  }}
  .insight-card.warning {{ border-left-color: #C45D3E; background: #FDF0ED; }}
  .insight-card.info {{ border-left-color: #D4A574; background: #FFF8F0; }}
  .insight-card.positive {{ border-left-color: #5B8C5A; background: #F0F7F0; }}
  .insight-icon {{ font-weight: 700; margin-right: 6px; }}
  .insight-card.warning .insight-icon {{ color: #C45D3E; }}
  .insight-card.info .insight-icon {{ color: #D4A574; }}
  .insight-card.positive .insight-icon {{ color: #5B8C5A; }}

  /* Feedback text list */
  .feedback-list {{ list-style: none; padding: 0; }}
  .feedback-list li {{
    padding: 10px 12px; margin-bottom: 8px; border-radius: 6px;
    background: var(--bg); font-size: 0.85rem; border: 1px solid var(--border);
  }}
  .feedback-list .meta {{ font-size: 0.75rem; color: var(--text-muted); margin-top: 4px; }}

  /* Nav tabs */
  .tabs {{ display: flex; gap: 4px; margin-bottom: 24px; flex-wrap: wrap; }}
  .tabs button {{
    padding: 8px 16px; border: 1px solid var(--border); background: var(--surface);
    border-radius: 6px; cursor: pointer; font-size: 0.85rem; color: var(--text-muted);
    transition: all 0.15s;
  }}
  .tabs button:hover {{ border-color: var(--primary-light); }}
  .tabs button.active {{ background: var(--primary); color: white; border-color: var(--primary); }}
  .tab-content {{ display: none; }}
  .tab-content.active {{ display: block; }}
</style>
</head>
<body>

<header>
  <h1>What's That Bug &mdash; Analytics</h1>
  <span class="date-range">{start_date} &rarr; {end_date}</span>
</header>

<div id="insights-panel" class="insights"></div>

<div id="summary-cards" class="cards"></div>

<nav class="tabs" id="tabs">
  <button class="active" data-tab="traffic">Traffic</button>
  <button data-tab="audience">Audience</button>
  <button data-tab="game">Game</button>
  <button data-tab="funnel">Funnel</button>
  <button data-tab="feedback">Player Feedback</button>
  <button data-tab="image-quality">Image Quality</button>
  <button data-tab="rounds">Round Analysis</button>
</nav>

<!-- Traffic tab -->
<div id="tab-traffic" class="tab-content active">
  <section>
    <h2>Pageviews &amp; Sessions</h2>
    <div class="chart-box"><canvas id="chart-pageviews"></canvas></div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Top Pages</h3>
        <div id="table-pages"></div>
      </div>
      <div class="chart-box">
        <h3>Top Referrers</h3>
        <div id="table-referrers"></div>
      </div>
    </div>
  </section>
</div>

<!-- Audience tab -->
<div id="tab-audience" class="tab-content">
  <section>
    <div class="grid-3">
      <div class="chart-box">
        <h3>Devices</h3>
        <canvas id="chart-devices"></canvas>
      </div>
      <div class="chart-box">
        <h3>Browsers</h3>
        <canvas id="chart-browsers"></canvas>
      </div>
      <div class="chart-box">
        <h3>Operating Systems</h3>
        <canvas id="chart-os"></canvas>
      </div>
    </div>
  </section>
  <section>
    <div class="chart-box">
      <h3>Top Countries</h3>
      <canvas id="chart-countries"></canvas>
    </div>
  </section>
</div>

<!-- Game tab -->
<div id="tab-game" class="tab-content">
  <section>
    <h2>Game Sessions</h2>
    <div class="chart-box"><canvas id="chart-game-sessions"></canvas></div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Sessions by Set</h3>
        <canvas id="chart-sets"></canvas>
      </div>
      <div class="chart-box">
        <h3>Average Score by Set</h3>
        <canvas id="chart-scores"></canvas>
      </div>
    </div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Daily Challenge</h3>
        <canvas id="chart-daily"></canvas>
      </div>
      <div class="chart-box">
        <h3>Event Breakdown</h3>
        <canvas id="chart-events"></canvas>
      </div>
    </div>
  </section>
</div>

<!-- Funnel tab -->
<div id="tab-funnel" class="tab-content">
  <section>
    <h2>Visitor &rarr; Player Funnel</h2>
    <div class="chart-box"><canvas id="chart-funnel"></canvas></div>
  </section>
</div>

<!-- Feedback tab -->
<div id="tab-feedback" class="tab-content">
  <section>
    <h2>Player Feedback</h2>
    <div id="feedback-summary-cards" class="cards"></div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Difficulty Rating Distribution</h3>
        <canvas id="chart-difficulty"></canvas>
      </div>
      <div class="chart-box">
        <h3>Would Play Again?</h3>
        <canvas id="chart-play-again"></canvas>
      </div>
    </div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Top Reported Observations (Bad Photo)</h3>
        <div id="table-bad-photos"></div>
      </div>
      <div class="chart-box">
        <h3>Free Text Feedback</h3>
        <div id="feedback-text-list"></div>
      </div>
    </div>
  </section>
</div>

<!-- Image Quality tab -->
<div id="tab-image-quality" class="tab-content">
  <section>
    <h2>Image Quality Analysis</h2>
    <div class="chart-box">
      <h3>Worst Observations by Miss Rate (min. 3 attempts)</h3>
      <div id="table-worst-obs"></div>
    </div>
  </section>
  <section>
    <div class="grid">
      <div class="chart-box">
        <h3>Top Confusion Pairs</h3>
        <div id="table-confusion"></div>
      </div>
      <div class="chart-box">
        <h3>Time Anomalies (avg &gt; 2x median)</h3>
        <div id="table-time-anomalies"></div>
      </div>
    </div>
  </section>
</div>

<!-- Round Analysis tab -->
<div id="tab-rounds" class="tab-content">
  <section>
    <h2>Round Analysis</h2>
    <div class="grid">
      <div class="chart-box">
        <h3>Drop-off by Round</h3>
        <canvas id="chart-dropoff"></canvas>
      </div>
      <div class="chart-box">
        <h3>Per-Round Difficulty Reactions</h3>
        <canvas id="chart-round-reactions"></canvas>
      </div>
    </div>
  </section>
  <section>
    <div class="chart-box">
      <h3>Round Accuracy by Set</h3>
      <div id="table-round-accuracy"></div>
    </div>
  </section>
</div>

<script>
// ---- Data injected by dump.py ----
const DATA = {data_json};

// ---- Palette ----
const COLORS = {{
  primary: '#C45D3E',
  primaryLight: '#E8A87C',
  accent: '#85603F',
  success: '#5B8C5A',
  muted: '#8B7355',
  chartPalette: ['#C45D3E','#E8A87C','#85603F','#5B8C5A','#D4A574','#A0522D','#CD853F','#DEB887','#BC8F8F','#F4A460'],
}};

const chartDefaults = {{
  responsive: true,
  maintainAspectRatio: true,
  plugins: {{
    legend: {{ labels: {{ font: {{ size: 11 }}, color: '#8B7355' }} }},
  }},
  scales: {{
    x: {{ ticks: {{ font: {{ size: 10 }}, color: '#8B7355' }}, grid: {{ display: false }} }},
    y: {{ ticks: {{ font: {{ size: 10 }}, color: '#8B7355' }}, grid: {{ color: '#E8DDD4' }} }},
  }},
}};

// ---- Tabs ----
document.querySelectorAll('.tabs button').forEach(btn => {{
  btn.addEventListener('click', () => {{
    document.querySelectorAll('.tabs button').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
    btn.classList.add('active');
    document.getElementById('tab-' + btn.dataset.tab).classList.add('active');
  }});
}});

// ---- Helpers ----
function get(key) {{ return DATA[key] || []; }}
function num(v) {{ return parseInt(v, 10) || 0; }}
function fmt(n) {{ return n.toLocaleString(); }}

function makeBarTable(containerId, rows, labelKey, valueKey, limit = 10) {{
  const el = document.getElementById(containerId);
  if (!rows.length) {{ el.innerHTML = '<p class="empty-state">No data</p>'; return; }}
  const top = rows.slice(0, limit);
  const maxVal = Math.max(...top.map(r => num(r[valueKey])));
  let html = '<table class="data-table"><thead><tr><th></th><th style="text-align:right"></th></tr></thead><tbody>';
  for (const r of top) {{
    const pct = maxVal > 0 ? (num(r[valueKey]) / maxVal * 100) : 0;
    const label = r[labelKey] || '(direct)';
    html += `<tr><td>${{label}}<br><span class="bar" style="width:${{pct}}%"></span></td><td style="text-align:right;white-space:nowrap">${{fmt(num(r[valueKey]))}}</td></tr>`;
  }}
  html += '</tbody></table>';
  el.innerHTML = html;
}}

function doughnutChart(canvasId, rows, labelKey, valueKey, limit = 8) {{
  const canvas = document.getElementById(canvasId);
  if (!rows.length) {{ canvas.parentElement.innerHTML += '<p class="empty-state">No data</p>'; return; }}
  const top = rows.slice(0, limit);
  new Chart(canvas, {{
    type: 'doughnut',
    data: {{
      labels: top.map(r => r[labelKey] || 'Unknown'),
      datasets: [{{ data: top.map(r => num(r[valueKey])), backgroundColor: COLORS.chartPalette }}],
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }}, color: '#8B7355', padding: 12 }} }} }},
    }},
  }});
}}

// ---- Summary Cards ----
(function() {{
  const stats = get('umami/stats')[0] || {{}};
  const game = DATA['game/aggregated'] || {{}};
  const cards = [
    {{ label: 'Pageviews', key: 'pageviews', change: 'pageviews_change' }},
    {{ label: 'Visitors', key: 'visitors', change: 'visitors_change' }},
    {{ label: 'Visits', key: 'visits', change: 'visits_change' }},
    {{ label: 'Bounce Rate', key: 'bounces', change: 'bounces_change', isBounce: true }},
  ];
  const container = document.getElementById('summary-cards');

  for (const c of cards) {{
    const val = num(stats[c.key]);
    const ch = num(stats[c.change]);
    let display = fmt(val);
    if (c.isBounce) {{
      const visits = num(stats['visits']);
      display = visits > 0 ? Math.round(val / visits * 100) + '%' : '—';
    }}
    const arrow = ch > 0 ? '&#9650;' : ch < 0 ? '&#9660;' : '';
    const cls = c.isBounce ? (ch <= 0 ? 'up' : 'down') : (ch >= 0 ? 'up' : 'down');
    container.innerHTML += `<div class="card"><div class="label">${{c.label}}</div><div class="value">${{display}}</div><div class="change ${{cls}}">${{arrow}} ${{Math.abs(ch)}}% vs prev period</div></div>`;
  }}

  if (game.total_sessions) {{
    container.innerHTML += `<div class="card"><div class="label">Game Sessions</div><div class="value">${{fmt(game.total_sessions)}}</div><div class="change" style="color:var(--text-muted)">from Google Sheets</div></div>`;
  }}
}})();

// ---- Traffic ----
(function() {{
  const pv = get('umami/pageviews');
  if (!pv.length) return;
  new Chart(document.getElementById('chart-pageviews'), {{
    type: 'line',
    data: {{
      labels: pv.map(r => r.date),
      datasets: [
        {{ label: 'Pageviews', data: pv.map(r => num(r.pageviews)), borderColor: COLORS.primary, backgroundColor: COLORS.primary + '20', fill: true, tension: 0.3 }},
        {{ label: 'Sessions', data: pv.map(r => num(r.sessions)), borderColor: COLORS.accent, backgroundColor: COLORS.accent + '20', fill: true, tension: 0.3 }},
      ],
    }},
    options: chartDefaults,
  }});

  makeBarTable('table-pages', get('umami/pages'), 'name', 'count');
  makeBarTable('table-referrers', get('umami/referrers'), 'name', 'count');
}})();

// ---- Audience ----
(function() {{
  doughnutChart('chart-devices', get('umami/devices'), 'name', 'count');
  doughnutChart('chart-browsers', get('umami/browsers'), 'name', 'count');
  doughnutChart('chart-os', get('umami/os'), 'name', 'count');

  const countries = get('umami/countries').slice(0, 15);
  if (countries.length) {{
    new Chart(document.getElementById('chart-countries'), {{
      type: 'bar',
      data: {{
        labels: countries.map(r => r.name),
        datasets: [{{ data: countries.map(r => num(r.count)), backgroundColor: COLORS.chartPalette }}],
      }},
      options: {{ ...chartDefaults, indexAxis: 'y', plugins: {{ legend: {{ display: false }} }} }},
    }});
  }}
}})();

// ---- Game Engagement (uses pre-aggregated data) ----
(function() {{
  const game = DATA['game/aggregated'];
  if (!game) {{
    document.getElementById('tab-game').innerHTML = '<p class="empty-state">No game data. Provide --xlsx to include Google Sheets data.</p>';
    return;
  }}

  // Sessions over time (by day)
  const sessionsByDay = game.sessions_by_day || {{}};
  const sortedDays = Object.keys(sessionsByDay).sort();
  if (sortedDays.length) {{
    new Chart(document.getElementById('chart-game-sessions'), {{
      type: 'bar',
      data: {{
        labels: sortedDays,
        datasets: [{{ label: 'Game Sessions', data: sortedDays.map(d => sessionsByDay[d]), backgroundColor: COLORS.primaryLight }}],
      }},
      options: {{ ...chartDefaults, plugins: {{ legend: {{ display: false }} }} }},
    }});
  }}

  // Sessions by set
  const setEntries = Object.entries(game.sessions_by_set || {{}}).sort((a, b) => b[1] - a[1]);
  if (setEntries.length) {{
    new Chart(document.getElementById('chart-sets'), {{
      type: 'bar',
      data: {{
        labels: setEntries.map(e => e[0]),
        datasets: [{{ data: setEntries.map(e => e[1]), backgroundColor: COLORS.chartPalette }}],
      }},
      options: {{ ...chartDefaults, indexAxis: 'y', plugins: {{ legend: {{ display: false }} }} }},
    }});
  }}

  // Average score by set
  const scoreEntries = Object.entries(game.avg_scores_by_set || {{}}).sort((a, b) => b[1] - a[1]);
  if (scoreEntries.length) {{
    new Chart(document.getElementById('chart-scores'), {{
      type: 'bar',
      data: {{
        labels: scoreEntries.map(e => e[0]),
        datasets: [{{ data: scoreEntries.map(e => e[1]), backgroundColor: COLORS.chartPalette }}],
      }},
      options: {{ ...chartDefaults, indexAxis: 'y', plugins: {{ legend: {{ display: false }} }} }},
    }});
  }}

  // Daily challenge
  const dailyByDay = game.daily_by_day || {{}};
  const dailyDays = Object.keys(dailyByDay).sort();
  if (dailyDays.length) {{
    new Chart(document.getElementById('chart-daily'), {{
      type: 'bar',
      data: {{
        labels: dailyDays,
        datasets: [
          {{ label: 'Starts', data: dailyDays.map(d => dailyByDay[d].starts), backgroundColor: COLORS.primaryLight }},
          {{ label: 'Completes', data: dailyDays.map(d => dailyByDay[d].completes), backgroundColor: COLORS.success }},
        ],
      }},
      options: chartDefaults,
    }});
  }}

  // Event type breakdown
  const typeEntries = Object.entries(game.event_types || {{}}).sort((a, b) => b[1] - a[1]);
  if (typeEntries.length) {{
    new Chart(document.getElementById('chart-events'), {{
      type: 'doughnut',
      data: {{
        labels: typeEntries.map(e => e[0]),
        datasets: [{{ data: typeEntries.map(e => e[1]), backgroundColor: COLORS.chartPalette }}],
      }},
      options: {{
        responsive: true,
        plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 10 }}, color: '#8B7355', padding: 8 }} }} }},
      }},
    }});
  }}
}})();

// ---- Funnel (uses pre-aggregated data) ----
(function() {{
  const stats = get('umami/stats')[0] || {{}};
  const game = DATA['game/aggregated'] || {{}};
  const lbCount = (DATA['game/leaderboard_count'] || [{{}}])[0];

  const visitors = num(stats['visitors']);
  const sessions = game.total_sessions || 0;
  const completes = game.total_completes || 0;
  const dailyPlays = game.total_daily_starts || 0;
  const lbEntries = num(lbCount.count);

  const steps = [
    {{ label: 'Visitors', value: visitors }},
    {{ label: 'Game Sessions', value: sessions }},
    {{ label: 'Completed Games', value: completes }},
    {{ label: 'Daily Challenge Plays', value: dailyPlays }},
    {{ label: 'Leaderboard Entries', value: lbEntries }},
  ].filter(s => s.value > 0);

  if (steps.length < 2) {{
    document.getElementById('tab-funnel').innerHTML = '<p class="empty-state">Not enough data for funnel. Need both Umami and Sheets data.</p>';
    return;
  }}

  new Chart(document.getElementById('chart-funnel'), {{
    type: 'bar',
    data: {{
      labels: steps.map(s => s.label),
      datasets: [{{
        data: steps.map(s => s.value),
        backgroundColor: steps.map((_, i) => COLORS.chartPalette[i % COLORS.chartPalette.length]),
      }}],
    }},
    options: {{
      ...chartDefaults,
      plugins: {{
        legend: {{ display: false }},
        tooltip: {{
          callbacks: {{
            afterLabel: function(ctx) {{
              if (ctx.dataIndex === 0) return '';
              const prev = steps[ctx.dataIndex - 1].value;
              const curr = steps[ctx.dataIndex].value;
              return prev > 0 ? `${{Math.round(curr / prev * 100)}}% of ${{steps[ctx.dataIndex - 1].label}}` : '';
            }},
          }},
        }},
      }},
    }},
  }});
}})();

// ---- Insights Panel ----
(function() {{
  const game = DATA['game/aggregated'] || {{}};
  const insights = game.insights || [];
  const panel = document.getElementById('insights-panel');
  if (!insights.length) return;

  const icons = {{ warning: '!', info: '?', positive: '+' }};
  let html = '<h2>Insights</h2><div class="insight-cards">';
  for (const ins of insights) {{
    const cls = ins.type || 'info';
    const icon = icons[cls] || '?';
    html += `<div class="insight-card ${{cls}}"><span class="insight-icon">${{icon}}</span>${{ins.text}}</div>`;
  }}
  html += '</div>';
  panel.innerHTML = html;
}})();

// ---- Player Feedback Tab ----
(function() {{
  const game = DATA['game/aggregated'];
  if (!game) return;

  // Summary cards
  const container = document.getElementById('feedback-summary-cards');
  const totalFeedback = Object.values(game.difficulty_ratings || {{}}).reduce((a, b) => a + b, 0);
  const playAgain = game.play_again_intent || {{}};
  const totalPA = Object.values(playAgain).reduce((a, b) => a + b, 0);
  const yesPA = playAgain.yes || 0;

  // Compute weighted average difficulty
  let weightedSum = 0, weightedCount = 0;
  for (const [rating, count] of Object.entries(game.difficulty_ratings || {{}})) {{
    weightedSum += parseInt(rating) * count;
    weightedCount += count;
  }}
  const avgDifficulty = weightedCount > 0 ? (weightedSum / weightedCount).toFixed(1) : '—';
  const playAgainPct = totalPA > 0 ? Math.round(yesPA / totalPA * 100) + '%' : '—';

  container.innerHTML = `
    <div class="card"><div class="label">Feedback Submissions</div><div class="value">${{fmt(totalFeedback)}}</div></div>
    <div class="card"><div class="label">Avg Difficulty</div><div class="value">${{avgDifficulty}}</div><div class="change" style="color:var(--text-muted)">out of 5</div></div>
    <div class="card"><div class="label">Would Play Again</div><div class="value">${{playAgainPct}}</div><div class="change" style="color:var(--text-muted)">${{yesPA}} of ${{totalPA}} responses</div></div>
    <div class="card"><div class="label">Bad Photo Reports</div><div class="value">${{fmt(Object.keys(game.bad_photo_reports || {{}}).length)}}</div><div class="change" style="color:var(--text-muted)">unique observations</div></div>
  `;

  // Difficulty rating distribution (bar chart)
  const ratings = game.difficulty_ratings || {{}};
  const ratingLabels = ['1', '2', '3', '4', '5'];
  const ratingData = ratingLabels.map(r => ratings[r] || 0);
  if (ratingData.some(v => v > 0)) {{
    new Chart(document.getElementById('chart-difficulty'), {{
      type: 'bar',
      data: {{
        labels: ratingLabels.map(r => r + ' star' + (r === '1' ? '' : 's')),
        datasets: [{{ data: ratingData, backgroundColor: ['#5B8C5A', '#85603F', '#E8A87C', '#D4A574', '#C45D3E'] }}],
      }},
      options: {{ ...chartDefaults, plugins: {{ legend: {{ display: false }} }} }},
    }});
  }}

  // Play again (doughnut)
  const paLabels = Object.keys(playAgain);
  const paData = paLabels.map(k => playAgain[k]);
  if (paLabels.length) {{
    new Chart(document.getElementById('chart-play-again'), {{
      type: 'doughnut',
      data: {{
        labels: paLabels.map(l => l.charAt(0).toUpperCase() + l.slice(1)),
        datasets: [{{ data: paData, backgroundColor: ['#5B8C5A', '#E8A87C', '#C45D3E'] }}],
      }},
      options: {{
        responsive: true,
        plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }}, color: '#8B7355', padding: 12 }} }} }},
      }},
    }});
  }}

  // Bad photo reports table
  const badPhotos = game.bad_photo_reports || {{}};
  const badPhotoEntries = Object.entries(badPhotos)
    .map(([id, info]) => ({{ obs_id: id, count: info.count, species: info.species }}))
    .sort((a, b) => b.count - a.count);
  const bpEl = document.getElementById('table-bad-photos');
  if (badPhotoEntries.length) {{
    let html = '<table class="data-table"><thead><tr><th>Observation ID</th><th>Species</th><th style="text-align:right">Reports</th></tr></thead><tbody>';
    for (const bp of badPhotoEntries.slice(0, 50)) {{
      html += `<tr><td>${{bp.obs_id}}</td><td>${{bp.species}}</td><td style="text-align:right">${{bp.count}}</td></tr>`;
    }}
    html += '</tbody></table>';
    if (badPhotoEntries.length > 50) html += `<p style="color:var(--text-muted);font-size:0.8rem;margin-top:8px">Showing 50 of ${{badPhotoEntries.length}} observations</p>`;
    bpEl.innerHTML = html;
  }} else {{
    bpEl.innerHTML = '<p class="empty-state">No bad photo reports</p>';
  }}

  // Free text feedback
  const texts = game.feedback_texts || [];
  const ftEl = document.getElementById('feedback-text-list');
  if (texts.length) {{
    let html = '<ul class="feedback-list">';
    for (const fb of texts.slice(0, 50)) {{
      html += `<li>${{fb.text}}<div class="meta">Difficulty: ${{fb.difficulty || '—'}} | Play again: ${{fb.play_again || '—'}} | ${{fb.timestamp || ''}}</div></li>`;
    }}
    html += '</ul>';
    if (texts.length > 50) html += `<p style="color:var(--text-muted);font-size:0.8rem">Showing 50 of ${{texts.length}} entries</p>`;
    ftEl.innerHTML = html;
  }} else {{
    ftEl.innerHTML = '<p class="empty-state">No free text feedback</p>';
  }}
}})();

// ---- Image Quality Tab ----
(function() {{
  const game = DATA['game/aggregated'];
  if (!game) return;

  // Worst observations table
  const worst = game.worst_observations || [];
  const woEl = document.getElementById('table-worst-obs');
  if (worst.length) {{
    let html = '<table class="data-table"><thead><tr><th>Obs ID</th><th>Species</th><th>Set</th><th style="text-align:right">Miss Rate</th><th style="text-align:right">Attempts</th><th style="text-align:right">Avg Time</th><th style="text-align:right">Bad Photo</th></tr></thead><tbody>';
    for (const o of worst) {{
      const timeStr = o.avg_time_ms > 0 ? (o.avg_time_ms / 1000).toFixed(1) + 's' : '—';
      const missColor = o.miss_rate > 80 ? 'color:#C45D3E;font-weight:600' : o.miss_rate > 50 ? 'color:#D4A574' : '';
      html += `<tr><td>${{o.obs_id}}</td><td>${{o.species}}</td><td>${{o.set}}</td><td style="text-align:right;${{missColor}}">${{o.miss_rate}}%</td><td style="text-align:right">${{o.attempts}}</td><td style="text-align:right">${{timeStr}}</td><td style="text-align:right">${{o.bad_photo_reports || 0}}</td></tr>`;
    }}
    html += '</tbody></table>';
    woEl.innerHTML = html;
  }} else {{
    woEl.innerHTML = '<p class="empty-state">Not enough data (need observations with 3+ attempts)</p>';
  }}

  // Confusion pairs table
  const confusion = game.confusion_pairs || [];
  const cpEl = document.getElementById('table-confusion');
  if (confusion.length) {{
    let html = '<table class="data-table"><thead><tr><th>Correct Answer</th><th>Player Picked</th><th style="text-align:right">Count</th></tr></thead><tbody>';
    for (const c of confusion) {{
      html += `<tr><td>${{c.correct}}</td><td>${{c.picked}}</td><td style="text-align:right">${{c.count}}</td></tr>`;
    }}
    html += '</tbody></table>';
    cpEl.innerHTML = html;
  }} else {{
    cpEl.innerHTML = '<p class="empty-state">No confusion pairs found</p>';
  }}

  // Time anomalies table
  const anomalies = game.time_anomalies || [];
  const taEl = document.getElementById('table-time-anomalies');
  if (anomalies.length) {{
    const medianStr = game.global_median_time_ms ? (game.global_median_time_ms / 1000).toFixed(1) + 's' : '—';
    let html = `<p style="font-size:0.8rem;color:var(--text-muted);margin-bottom:8px">Global median response time: ${{medianStr}}</p>`;
    html += '<table class="data-table"><thead><tr><th>Obs ID</th><th>Species</th><th style="text-align:right">Avg Time</th><th style="text-align:right">Median Ratio</th><th style="text-align:right">Attempts</th></tr></thead><tbody>';
    for (const a of anomalies.slice(0, 50)) {{
      html += `<tr><td>${{a.obs_id}}</td><td>${{a.species}}</td><td style="text-align:right">${{(a.avg_time_ms / 1000).toFixed(1)}}s</td><td style="text-align:right">${{a.median_ratio}}x</td><td style="text-align:right">${{a.attempts}}</td></tr>`;
    }}
    html += '</tbody></table>';
    if (anomalies.length > 50) html += `<p style="color:var(--text-muted);font-size:0.8rem;margin-top:8px">Showing 50 of ${{anomalies.length}} anomalies</p>`;
    taEl.innerHTML = html;
  }} else {{
    taEl.innerHTML = '<p class="empty-state">No time anomalies detected</p>';
  }}
}})();

// ---- Round Analysis Tab ----
(function() {{
  const game = DATA['game/aggregated'];
  if (!game) return;

  // Drop-off by round (bar chart)
  const dropoff = game.dropoff_by_round || {{}};
  const dropoffRounds = Object.keys(dropoff).sort((a, b) => parseInt(a) - parseInt(b));
  if (dropoffRounds.length) {{
    new Chart(document.getElementById('chart-dropoff'), {{
      type: 'bar',
      data: {{
        labels: dropoffRounds.map(r => 'Round ' + r),
        datasets: [{{
          label: '% of sessions reaching this round',
          data: dropoffRounds.map(r => dropoff[r]),
          backgroundColor: dropoffRounds.map((r, i) => {{
            const pct = dropoff[r];
            return pct > 70 ? '#5B8C5A' : pct > 40 ? '#E8A87C' : '#C45D3E';
          }}),
        }}],
      }},
      options: {{
        ...chartDefaults,
        plugins: {{ legend: {{ display: false }} }},
        scales: {{
          ...chartDefaults.scales,
          y: {{ ...chartDefaults.scales.y, max: 100, ticks: {{ ...chartDefaults.scales.y.ticks, callback: v => v + '%' }} }},
        }},
      }},
    }});
  }}

  // Per-round difficulty reactions (stacked bar)
  const reactions = game.round_reactions || {{}};
  const reactionRounds = Object.keys(reactions).sort((a, b) => parseInt(a) - parseInt(b));
  if (reactionRounds.length) {{
    new Chart(document.getElementById('chart-round-reactions'), {{
      type: 'bar',
      data: {{
        labels: reactionRounds.map(r => 'Round ' + r),
        datasets: [
          {{
            label: 'Too Easy',
            data: reactionRounds.map(r => reactions[r].too_easy || 0),
            backgroundColor: '#5B8C5A',
          }},
          {{
            label: 'Just Right',
            data: reactionRounds.map(r => reactions[r].just_right || 0),
            backgroundColor: '#E8A87C',
          }},
          {{
            label: 'Too Hard',
            data: reactionRounds.map(r => reactions[r].too_hard || 0),
            backgroundColor: '#C45D3E',
          }},
        ],
      }},
      options: {{
        ...chartDefaults,
        plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }}, color: '#8B7355', padding: 12 }} }} }},
        scales: {{
          ...chartDefaults.scales,
          x: {{ ...chartDefaults.scales.x, stacked: true }},
          y: {{ ...chartDefaults.scales.y, stacked: true }},
        }},
      }},
    }});
  }}

  // Round accuracy by set (table)
  const roundAcc = game.round_accuracy_by_set || {{}};
  const sets = Object.keys(roundAcc).sort();
  const raEl = document.getElementById('table-round-accuracy');
  if (sets.length) {{
    // Find all round numbers across all sets
    const allRounds = new Set();
    for (const s of sets) {{
      for (const r of Object.keys(roundAcc[s])) allRounds.add(r);
    }}
    const sortedRounds = [...allRounds].sort((a, b) => parseInt(a) - parseInt(b));

    let html = '<table class="data-table"><thead><tr><th>Set</th>';
    for (const r of sortedRounds) html += `<th style="text-align:center">R${{r}}</th>`;
    html += '</tr></thead><tbody>';
    for (const s of sets) {{
      html += `<tr><td>${{s}}</td>`;
      for (const r of sortedRounds) {{
        const data = (roundAcc[s] || {{}})[r];
        if (data && data.total > 0) {{
          const pct = Math.round(data.correct / data.total * 100);
          const color = pct >= 70 ? '#5B8C5A' : pct >= 40 ? '#D4A574' : '#C45D3E';
          html += `<td style="text-align:center;color:${{color}}">${{pct}}%<br><span style="font-size:0.7rem;color:var(--text-muted)">n=${{data.total}}</span></td>`;
        }} else {{
          html += '<td style="text-align:center;color:var(--text-muted)">—</td>';
        }}
      }}
      html += '</tr>';
    }}
    html += '</tbody></table>';
    raEl.innerHTML = html;
  }} else {{
    raEl.innerHTML = '<p class="empty-state">No round accuracy data</p>';
  }}
}})();

</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="What's That Bug — Analytics Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--days", type=int, default=30, help="Number of days to pull (default: 30)")
    parser.add_argument("--start", type=str, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", type=str, help="End date (YYYY-MM-DD, default: today)")
    parser.add_argument("--xlsx", type=str, help="Path to Google Sheets xlsx export")
    parser.add_argument("--dump-only", action="store_true", help="Only dump CSVs, skip dashboard")
    parser.add_argument("--dashboard-only", action="store_true", help="Only generate dashboard from existing CSVs")
    args = parser.parse_args()

    load_env()

    # Date range
    if args.start:
        start_dt = datetime.strptime(args.start, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    else:
        start_dt = datetime.now(timezone.utc) - timedelta(days=args.days)
    if args.end:
        end_dt = datetime.strptime(args.end, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    else:
        end_dt = datetime.now(timezone.utc)

    start_ms = int(start_dt.timestamp() * 1000)
    end_ms = int(end_dt.timestamp() * 1000)
    start_str = start_dt.strftime("%Y-%m-%d")
    end_str = end_dt.strftime("%Y-%m-%d")

    print(f"Date range: {start_str} -> {end_str}")

    if not args.dashboard_only:
        # Umami dump
        api_key = os.environ.get("UMAMI_API_KEY", "")
        website_id = os.environ.get("UMAMI_WEBSITE_ID", "")
        if api_key and website_id:
            dump_umami(api_key, website_id, start_ms, end_ms)
        else:
            print("\n--- Umami Cloud ---")
            print("  Skipped: UMAMI_API_KEY or UMAMI_WEBSITE_ID not set in .env")

        # Sheets dump
        xlsx_path = None
        if args.xlsx:
            xlsx_path = Path(args.xlsx)
        else:
            # Auto-detect xlsx in project root
            project_root = SCRIPT_DIR.parent
            candidates = sorted(project_root.glob("What's That Bug*.xlsx"))
            if candidates:
                xlsx_path = candidates[-1]  # most recent
                print(f"\n  Auto-detected xlsx: {xlsx_path.name}")

        if xlsx_path:
            dump_sheets(xlsx_path)
        else:
            print("\n--- Google Sheets ---")
            print("  Skipped: no xlsx file found. Use --xlsx or place the export in the project root.")

    # Dashboard
    if not args.dump_only:
        generate_dashboard(start_str, end_str)

    print("\nDone!")


if __name__ == "__main__":
    main()
